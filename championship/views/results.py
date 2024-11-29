# Copyright 2024 Leonin League
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import csv
import io
import logging
import re
from collections import Counter
from typing import Iterable
from zipfile import BadZipFile

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ImproperlyConfigured
from django.db import transaction
from django.http import HttpResponseForbidden, HttpResponseRedirect
from django.shortcuts import get_object_or_404
from django.urls import reverse
from django.views.generic.edit import FormView, UpdateView

import requests
from openpyxl import load_workbook

from championship.forms import (
    AddTop8ResultsForm,
    FileImporterForm,
    ImporterSelectionForm,
    LinkImporterForm,
    ManualUploadMetadataForm,
    ResultForm,
    ResultsDeleteForm,
    ResultsFormset,
)
from championship.models import Event, Player, Result
from championship.parsers import (
    aetherhub,
    challonge,
    eventlink,
    excel_csv_parser,
    melee,
    mtgevent,
    spicerack,
)
from championship.parsers.general_parser_functions import parse_record, record_to_points
from championship.parsers.parse_result import ParseResult
from championship.tournament_valid import (
    TooManyPointsForPlayerError,
    TooManyPointsForTop8Error,
    TooManyPointsInTotalError,
    get_max_round_error_message,
    validate_standings,
)
from championship.views.base import CustomDeleteView


def validate_standings_and_show_error(request, standings, category):
    """
    Validates the standings for a given category and sends a given error message to the UI.

    Args:
        standings (list): A list of tuples containing player names and their respective points.
        category (Event.Category): The category of the event.

    Returns:
        True: When an error html should be rendered for the user.
    """
    try:
        validate_standings(standings, category)
    except (
        TooManyPointsForPlayerError,
        TooManyPointsInTotalError,
        TooManyPointsForTop8Error,
    ) as e:
        if category == Event.Category.REGULAR:
            error_message = f"{e.ui_error_message()} You're trying to upload a SUL Regular event with more than 6 Swiss rounds. Please contact us at {settings.PUBLIC_CONTACT_EMAIL}!"
        else:
            error_message = f"{e.ui_error_message()} {get_max_round_error_message(category, standings)} Please use the standings of the last Swiss round!"
        messages.error(request, error_message)
        return True
    return False


class CreateResultsView(FormView):
    """Generic view that handles the logic for importing results.

    This view encapsulate all the code that should be website-independent for
    importing results, such as validating a form, looking for events, etc.

    To use this class, the user simply needs to implement the get_results
    method, which contains all the parsing logic. See
    CreateAetherhubResultsView for an example.
    """

    template_name = "championship/create_results.html"

    def get_form_kwargs(self):
        k = super().get_form_kwargs()
        k["user"] = self.request.user
        return k

    def get_results(self, form):
        """This function must be implemented to actually parse results.

        It returns a list of standings, or None if parsing was not succesful.
        """
        raise ImproperlyConfigured("No parser implemented")

    @transaction.atomic
    def form_valid(self, form):
        """Processes a succesful result creation form.

        This part of the processing is generic and does not depend on which
        website we use for parsing.
        """
        # From here we can assume that the event exists and is owned by
        # this user, as otherwise the form validation will not accept it.
        self.event = form.cleaned_data["event"]
        standings = self.get_results(form)

        if not standings:
            return self.form_invalid(form)

        # Sometimes the webpages or users don't sort the standings correctly. Hence we should sort as a precaution.
        standings.sort(key=lambda pr: pr.points, reverse=True)

        # Check that the records amount to the same amount of match points
        for parse_result in standings:
            (w, l, d) = parse_result.record
            points = parse_result.points
            if points != w * 3 + d:
                messages.error(
                    self.request,
                    f"""The record of {parse_result.name} does not add up to the match points. Please send us
                    the results link or file via email to {settings.PUBLIC_CONTACT_EMAIL}""",
                )
                return self.form_invalid(form)

        if self.event.results_validation_enabled and validate_standings_and_show_error(
            self.request,
            [(pr.name, pr.points, pr.record) for pr in standings],
            self.event.category,
        ):
            return self.form_invalid(form)

        results_to_create = []
        for i, parse_result in enumerate(standings):
            (w, l, d) = parse_result.record
            player, created = Player.objects.get_or_create_by_name(parse_result.name)

            results_to_create.append(
                Result(
                    points=parse_result.points,
                    player=player,
                    event=self.event,
                    ranking=i + 1,
                    win_count=w,
                    loss_count=l,
                    draw_count=d,
                    decklist_url=parse_result.decklist_url or "",
                    deck_name=parse_result.deck_name or "",
                )
            )

        Result.objects.bulk_create(results_to_create)

        return super().form_valid(form)

    def get_success_url(self):
        return self.event.get_absolute_url()


class ManualResultsView(LoginRequiredMixin, CreateResultsView):
    template_name = "championship/create_results_manual.html"

    def get_context_data(self, formset=None, metadata_form=None):
        if not formset:
            formset = ResultsFormset()

        if not metadata_form:
            metadata_form = ManualUploadMetadataForm(user=self.request.user)

        players = Player.leaderboard_objects.all()

        return {
            "metadata_form": metadata_form,
            "formset": formset,
            "players": players,
        }

    def form_invalid(self, form):
        context = self.get_context_data(metadata_form=form, formset=self.formset)
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        self.formset = ResultsFormset(request.POST)
        metadata_form = ManualUploadMetadataForm(user=request.user, data=request.POST)

        if not self.formset.is_valid() or not metadata_form.is_valid():
            return self.form_invalid(metadata_form)
        return self.form_valid(metadata_form)

    def get_results(self, form):
        standings = []
        for result in self.formset.cleaned_data:
            name = result.get("name")
            record = result.get("points")
            if name and record:
                pr = ParseResult(
                    name=name,
                    points=record_to_points(record),
                    record=parse_record(record),
                )
                standings.append(pr)
        return standings


class CreateLinkParserResultsView(LoginRequiredMixin, CreateResultsView):
    form_class = LinkImporterForm
    help_text: str
    placeholder: str

    def clean_url(self, url):
        raise NotImplementedError("This method has not been implemented yet.")

    def extract_standings_from_page(self, text: str) -> Iterable[tuple[str, int]]:
        raise NotImplementedError("No parser yet")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({"help_text": self.help_text, "placeholder": self.placeholder})
        return kwargs

    def validate_response(self, response: requests.Response):
        """Validates the response and returns True if it is valid."""
        return True

    def get_results(self, form):
        url = form.cleaned_data["url"]
        url = self.clean_url(url)
        if not url:
            messages.error(self.request, "Wrong url format.")
            return
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                "Accept-Language": "en-US,en;q=0.9",
                "Connection": "keep-alive",
                "Cache-conrol": "max-age=0",
                "Sec-Fetch-Dest": "document",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Site": "none",
                "Sec-Fetch-User": "?1",
                "Upgrade-Insecure-Requests": "1",
            }
            response = requests.get(url, headers=headers)
            response.raise_for_status()

            if not self.validate_response(response):
                return

            return list(self.extract_standings_from_page(response.content.decode()))
        except Exception as e:
            logging.exception("Could not fetch standings")
            message = "Could not fetch standings."
            if hasattr(e, "ui_error_message"):
                message = e.ui_error_message
            messages.error(self.request, message)


class CreateFileParserResultsView(LoginRequiredMixin, CreateResultsView):
    form_class = FileImporterForm
    help_text = "The file that contains the standings."

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({"help_text": self.help_text})
        return kwargs


class CreateHTMLParserResultsView(CreateFileParserResultsView):
    help_text = (
        "The standings file saved as a web page (.html). "
        + "Go to the standings page of the last swiss round, then press Ctrl+S and save."
    )
    error_text = (
        "Error: Could not parse standings file. Did you upload it as HTML? "
        + "You can get a HTML file by going to the standings of the last swiss round and pressing Ctrl+S."
    )

    def extract_standings_from_page(self, text: str) -> Iterable[tuple[str, int]]:
        raise NotImplementedError("No parser yet")

    def get_results(self, form):
        try:
            text = "".join(s.decode() for s in self.request.FILES["standings"].chunks())
            return self.extract_standings_from_page(text)
        except Exception:
            logging.exception("Could not parse page")
            messages.error(self.request, self.error_text)


class AetherhubResultsView(CreateLinkParserResultsView):
    help_text = (
        "Link to your tournament. Make sure it is a public and finished tournament."
    )
    placeholder = "https://aetherhub.com/Tourney/RoundTourney/123456"

    def extract_standings_from_page(self, text):
        return aetherhub.parse_standings_page(text)

    def clean_url(self, url):
        """Normalizes the given tournament url to point to the RoundTourney page."""
        url_re = r"https://aetherhub.com/Tourney/[a-zA-Z]+/(\d+)"
        tourney = re.match(url_re, url)
        if tourney:
            return f"https://aetherhub.com/Tourney/RoundTourney/{tourney.group(1)}"

    def validate_response(self, response):
        if response.history and response.history[0].status_code == 302:
            messages.error(
                self.request,
                "The tournament was not found. Make sure it is a public tournament.",
            )
            return False
        if (
            response.status_code == 200
            and "Finished: " not in response.content.decode()
        ):
            messages.error(
                self.request,
                "The tournament is not finished. Please enter the results for all rounds and end the tournament.",
            )
            return False
        return True


# Curently disabled as challonge flags our requests as spam
class ChallongeLinkResultsView(CreateLinkParserResultsView):
    help_text = (
        "Link to your tournament. Make sure the tournament system is Swiss rounds."
    )
    placeholder = "https://challonge.com/de/rk6vluaa"

    def extract_standings_from_page(self, text):
        return challonge.parse_standings_page(text)

    def clean_url(self, url):
        try:
            return challonge.clean_url(url)
        except ValueError:
            logging.exception("Could not clean Challonge URL")


class ChallongeHtmlResultsView(CreateHTMLParserResultsView):
    def extract_standings_from_page(self, text):
        return challonge.parse_standings_page(text)


class EventlinkResultsView(CreateHTMLParserResultsView):
    def extract_standings_from_page(self, text):
        return eventlink.parse_standings_page(text)


class MtgEventResultsView(CreateHTMLParserResultsView):
    def extract_standings_from_page(self, text):
        return mtgevent.parse_standings_page(text)


class ExcelCsvResultsView(CreateFileParserResultsView):
    help_text = (
        "Upload an Excel (.xlsx) or CSV (.csv) file. The headers of the columns need to be named in a specific way: "
        + "PLAYER_NAME for the column with the name of the player. "
        + "RECORD for the record of the player. "
        + "You can also use the column MATCH_POINTS if you only have the match points and no record."
    )

    def _read_excel_or_csv_rows(self):
        file_buffer = io.BytesIO(self.request.FILES["standings"].read())

        try:
            workbook = load_workbook(file_buffer, data_only=True)
            sheet = workbook.active

            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([cell for cell in row if cell is not None])

            return rows
        except BadZipFile:
            file_buffer.seek(0)
            try:
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(file_buffer.read(1024).decode()).delimiter
                file_buffer.seek(0)

                reader = csv.reader(
                    io.TextIOWrapper(file_buffer, encoding="utf-8"), delimiter=delimiter
                )
                rows = [row for row in reader]
                return rows
            except (csv.Error, UnicodeDecodeError):
                return None

    def get_results(self, form):
        error_text = "Error when reading the file. Did you upload a .xlsx or .csv file with the headers of the columns named PLAYER_NAME and RECORD (or MATCH_POINTS)?"
        rows = self._read_excel_or_csv_rows()
        if rows is None:
            logging.exception("Could not parse file as Excel or CSV")
            messages.error(self.request, error_text)
            return
        try:
            return excel_csv_parser.parse_standings_page(rows)
        except Exception as e:
            logging.exception("Error parsing dataframe")
            if hasattr(e, "ui_error_message"):
                error_text = e.ui_error_message
            else:
                raise e
        messages.error(self.request, error_text)


class MeleeResultsView(CreateFileParserResultsView):
    help_text = "Upload the standings exported by clicking on 'Export All Standings' in Melee's Tournament Controller's Standings page (.csv file)."

    def get_results(self, form):
        text = "".join(s.decode() for s in self.request.FILES["standings"].chunks())
        return melee.parse_standings(text)


class SpicerackResultsView(CreateLinkParserResultsView):
    help_text = "Link to your tournament. Make sure that all Swiss rounds are finished."
    placeholder = "https://spicerack.gg/admin/events/1234567890"

    def extract_standings_from_page(self, text):
        return

    def get_results(self, form):
        url = form.cleaned_data["url"]
        event_id = spicerack.extract_event_id_from_url(url)
        if not event_id:
            messages.error(self.request, "Wrong url format.")
            return
        try:
            response = requests.get(
                f"https://hydra.spicerack.gg/api/magic-events/{event_id}/get_all_rounds/"
            )
            response.raise_for_status()

            round = spicerack.parse_rounds_json(response.json())
            round_id = round["id"]

            response = requests.get(
                f"https://hydra.spicerack.gg/api/tournament-rounds/{round_id}/include_all_standings/"
            )
            response.raise_for_status()

            return spicerack.parse_standings_json(
                response.json(), round["round_number"]
            )
        except Exception as e:
            logging.exception("Could not fetch standings")
            message = "Could not fetch standings."
            if hasattr(e, "ui_error_message"):
                message = e.ui_error_message
            messages.error(self.request, message)


class ChooseUploaderView(LoginRequiredMixin, FormView):
    template_name = "championship/create_results.html"
    form_class = ImporterSelectionForm

    def form_valid(self, form):
        from championship.importers import IMPORTER_LIST

        urls_for_type = {
            parser.name.upper(): reverse(parser.view_name) for parser in IMPORTER_LIST
        }
        return HttpResponseRedirect(urls_for_type[form.cleaned_data["site"]])


class AddTop8ResultsView(LoginRequiredMixin, FormView):
    template_name = "championship/add_top8_results.html"
    form_class = AddTop8ResultsForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["event"] = get_object_or_404(
            Event, pk=self.kwargs["pk"], organizer__user=self.request.user
        )
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        self.event = Event.objects.get(id=self.kwargs["pk"])

        if not self.event.can_be_edited():
            messages.error(
                self.request, "Event too old to add or change the playoffs results."
            )
            return super().form_valid(form)

        FIELDS_TO_RESULTS = {
            "winner": Result.PlayoffResult.WINNER,
            "finalist": Result.PlayoffResult.FINALIST,
            "semi0": Result.PlayoffResult.SEMI_FINALIST,
            "semi1": Result.PlayoffResult.SEMI_FINALIST,
            "quarter0": Result.PlayoffResult.QUARTER_FINALIST,
            "quarter1": Result.PlayoffResult.QUARTER_FINALIST,
            "quarter2": Result.PlayoffResult.QUARTER_FINALIST,
            "quarter3": Result.PlayoffResult.QUARTER_FINALIST,
        }

        playoff_results_filled = [
            (form.cleaned_data[key], single_elim_result)
            for key, single_elim_result in FIELDS_TO_RESULTS.items()
            if form.cleaned_data[key] is not None
        ]

        if len(playoff_results_filled) not in [4, 8]:
            messages.error(self.request, "You need to fill in 4 or 8 playoff results.")
            return super().form_invalid(form)

        counter = Counter([epr for epr, _ in playoff_results_filled])
        duplicates = [result for result, count in counter.items() if count > 1]
        if duplicates:
            messages.error(
                self.request,
                f"Player '{duplicates[0].player.name}'has more than 1 result.",
            )
            return super().form_invalid(form)

        self.event.result_set.update(playoff_result=None)
        for event_player_result, single_elim_result in playoff_results_filled:
            event_player_result.playoff_result = single_elim_result
            event_player_result.save()

        return super().form_valid(form)

    def get_success_url(self):
        return self.event.get_absolute_url()


class ClearEventResultsView(LoginRequiredMixin, FormView):
    template_name = "championship/results_confirm_delete.html"
    form_class = ResultsDeleteForm

    def get_event(self) -> Event:
        return get_object_or_404(
            Event, id=self.kwargs["pk"], organizer__user=self.request.user
        )

    def get_success_url(self):
        return self.get_event().get_absolute_url()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["event"] = self.get_event()
        return ctx

    def form_valid(self, form):
        # No processing to do here, just delete results
        event = self.get_event()

        if event.can_be_edited():
            event.result_set.all().delete()
        else:
            messages.error(self.request, "Event too old to delete results.")

        return super().form_valid(form)


def update_ranking_order(event):
    """Updates the order of the ranking after a result has been updated."""
    results = Result.objects.filter(event=event)
    results = sorted(
        results, key=lambda r: (r.win_count, r.draw_count, -r.ranking), reverse=True
    )
    for i, result in enumerate(results):
        result.ranking = i + 1
        result.save()


class ResultUpdatePermissionMixin:
    """Contains the logic to check if the user is allowed to update a specific result."""

    def dispatch(self, request, *args, **kwargs):
        event = self.get_object().event
        organizer_allowed_to_edit = (
            request.user == event.organizer.user and event.can_be_edited()
        )
        if request.user.is_superuser or organizer_allowed_to_edit:
            return super().dispatch(request, *args, **kwargs)
        else:
            return HttpResponseForbidden()


class ResultUpdateView(ResultUpdatePermissionMixin, UpdateView):
    model = Result
    form_class = ResultForm
    template_name = "championship/update_result.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["players"] = Player.leaderboard_objects.all()
        return context

    @transaction.atomic
    def form_valid(self, form):
        old_player = self.get_object().player
        form.save()
        # Delete the old player if they have no results anymore
        results_old_player = Result.objects.filter(player=old_player)
        if not results_old_player:
            old_player.delete()
        update_ranking_order(form.instance.event)
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse("event_details", args=[self.object.event.id])


class SingleResultDeleteView(ResultUpdatePermissionMixin, CustomDeleteView):
    model = Result

    def get_success_url(self):
        return reverse("event_details", args=[self.object.event.id])

    def form_valid(self, form):
        update_ranking_order(self.object.event)

        self.object = self.get_object()
        self.object.delete()
        if not Result.objects.filter(player=self.object.player_id).count():
            self.object.player.delete()

        return HttpResponseRedirect(self.get_success_url())
